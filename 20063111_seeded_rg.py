import numpy as np
import cv2
import os
import glob
import openpyxl as pyxl
from matplotlib import pyplot as pt

img_list = []

#Get user input
def getList():
    img_dir_list = []
    validate = ['1093','4795','8438','10870']
    counter = 0
    while True:
        img_dir = input('Please enter dir to be included. (1093,4795,8438,10870)\nEnter "All" to select all folder\nEnter "Done" to stop selection: ')
       
        if img_dir in validate:
            if not img_dir in img_dir_list:
                img_dir_list.append(img_dir)
                counter += 1
                print("Directory " + str(img_dir) + " added.")
            else:
                print("Directory " + str(img_dir) + " already included.")
        elif img_dir == 'All' or img_dir == 'Done':
            if counter == 4 or img_dir == 'All':
                print("All directories chosen.")
                img_dir_list = validate.copy()
                break
        
            if img_dir == 'Done':
                if len(img_dir_list) == 0:
                    print("\nPlease enter atleast one directory")
                    print("------------------------------")
                else:
                    print("Directory: ")
                    print(img_dir_list)
                    print("selected.")
                    break
        else:
            print("\nPlease enter valid input/directory.")
            print("------------------------------")
    
    return img_dir_list

#Get all images files with respective dir, initialize dirs
def getImages():
    c_dir = os.getcwd()
    image_file_dirs = getList()
    segment_file_dir = os.path.join(c_dir,"segmentation_result")
    iou_file_dir = os.path.join(c_dir,"iou_score")
    mask_file_dir = os.path.join(c_dir,"mask")
    size_1093 = 0
    size_4795 = 0
    size_8438 = 0
    size_10870 = 0
    
    #Create segmentation directory
    for image_file_dir in image_file_dirs:
        if os.path.exists(os.path.join(segment_file_dir, image_file_dir)):
            # print("Directory already exist.")
            continue
        else:
            os.makedirs(os.path.join(segment_file_dir, image_file_dir))
            
    #Create CSV file
    if not os.path.exists(os.path.join(segment_file_dir, iou_file_dir)):
        os.makedirs(os.path.join(segment_file_dir, iou_file_dir))
        temp_dir = os.path.join(segment_file_dir, iou_file_dir)
        for img_catagory in image_file_dirs:
            if not os.path.exists(os.path.join(temp_dir, img_catagory + "_iou_score.xlsx")):
                wb = pyxl.Workbook()
                wb.save(os.path.join(temp_dir, img_catagory + "_iou_score.xlsx"))
    else:
        temp_dir = os.path.join(segment_file_dir, iou_file_dir)
        for img_catagory in image_file_dirs:
            if not os.path.exists(os.path.join(temp_dir, img_catagory + "_iou_score.xlsx")):
                wb = pyxl.Workbook()
                wb.save(os.path.join(temp_dir, img_catagory + "_iou_score.xlsx"))   
                
    #Get all image path
    for image_dir in image_file_dirs:
        for img_path in glob.glob(image_dir + "\*.jpg"):
            if image_dir == '1093':
                img_list.append(img_path)
                size_1093 += 1
            elif image_dir == '4795':
                img_list.append(img_path)
                size_4795 += 1
            elif image_dir == '8438':
                img_list.append(img_path)
                size_8438 += 1
            elif image_dir == '10870':
                img_list.append(img_path)
                size_10870 += 1
    
    return size_1093,size_4795,size_8438,size_10870

#Get ground truth image of each dataset
def getMask(mask_dir):
    #Get mask path
    for mask_dir in glob.glob("mask" + "\\" + mask_dir + ".png"):
        return mask_dir
        
#Seeded Region Growing
class Pixel(object):
    #Constructor for each coordinate
    def __init__(self,row,col):
        self.row = row
        self.col = col
    
    def getRow(self):
        return self.row
    
    def getCol(self):
        return self.col
    
def getConnect(opt):
    #Pixel Coordinate Connectivity SE
    connect_8 = [Pixel(-1, -1), Pixel(0, -1), Pixel(1, -1), Pixel(1, 0), Pixel(1, 1), Pixel(0, 1), Pixel(-1, 1), Pixel(-1, 0)]
    connect_4 = [Pixel(0, -1), Pixel(1, 0),Pixel(0, 1), Pixel(-1, 0)]
    #cv2 Dilation SE
    np_connect_8 = np.array([[1,1,1],[1,1,1],[1,1,1]],dtype=np.uint8)

    if opt == '8_px':
        return connect_8
    if opt == '4_px':
        return connect_4
    if opt == '8_np':
        return np_connect_8

def regionGrowing(img_parse,seeds,sky_thres,connect):
    image = img_parse.copy()
    connect_sE = getConnect(connect)
    ori_seed_list = seeds.copy()
    [nrow,ncol] = image.shape
    regionMap = np.zeros([nrow,ncol],dtype=np.uint8)
    region_num = 255
    
    sky_seed = seeds.copy()
    while True:    
        current_seed = sky_seed.pop(0)
        regionMap[current_seed.row,current_seed.col] = region_num
        
        for x in range(len(connect_sE)):
            neighbor_row = current_seed.row + connect_sE[x].row
            neighbor_col = current_seed.col + connect_sE[x].col
            if neighbor_row < 0 or neighbor_col < 0 or neighbor_row >= nrow or neighbor_col >= ncol:
                continue
            #Not Seed
            if not [Pixel(neighbor_row,neighbor_col)] in ori_seed_list:
                #Not Catagorized
                if regionMap[neighbor_row,neighbor_col] == 0:
                    #Not Exceed Thres
                    if abs(int(image[current_seed.row,current_seed.col]) - int(image[neighbor_row,neighbor_col])) < sky_thres and regionMap[neighbor_row,neighbor_col] == 0:
                        #Mark in region map as catagorized
                        regionMap[neighbor_row,neighbor_col] = region_num
                        #Add new grow seed  
                        sky_seed.append(Pixel(neighbor_row,neighbor_col))
            
        if len(sky_seed) == 0:
            print("Growing Complete")
            return regionMap
            break

def getSeed(image):
    sky_region = image[0:27,0:181]
    nrow,ncol,c = image.shape
    seedList = []
    day = True

    #seed 1 top left sky
    x = 25
    y = 25
    r,g,b = sky_region[x,y]
    
    # Night
    if r <= 120 and g <= 120 and b <= 120:
        seedList.append(Pixel(x,y))
        day = False
    #Day/Evening (Higher intensity values)
    else:
        seedList.append(Pixel(x,y))
        day = True
        
    #seed 2 top left sky(right side, after clock tower in 10870 image)
    x = 25
    y = 185
    seedList.append(Pixel(x,y))
        
    return day,seedList
    
size_1093,size_4795,size_8438,size_10870 = getImages()
img_set_size = 0
cell_row = 2
counter_1093 = 0
counter_4795 = 0
counter_8438 = 0
counter_10870 = 0
img_count = 0
for imgfile in img_list:
    if imgfile.find('1093') >= 0:
        if counter_1093 == 0:
            img_count = 0
            cell_row = 2
            counter_1093 += 1
        iou_file_name = '1093'
        img_set_size = size_1093
    if imgfile.find('4795') >= 0:
        if counter_4795 == 0:
            img_count = 0
            cell_row = 2
            counter_4795 += 1
        iou_file_name = '4795'
        img_set_size = size_4795
    if imgfile.find('8438') >= 0:
        if counter_8438 == 0:
            img_count = 0
            cell_row = 2
            counter_8438 += 1
        iou_file_name = '8438'
        img_set_size = size_8438
    if imgfile.find('10870') >= 0:
        if counter_10870 == 0:
            img_count = 0
            cell_row = 2
            counter_10870 += 1
        iou_file_name = '10870'
        img_set_size = size_10870

    mask_dir = getMask(iou_file_name)
        
    wb = pyxl.load_workbook('iou_score/' +iou_file_name + '_iou_score.xlsx')
    ws = wb.active
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 40
    
    ws['A1'] = iou_file_name + ' Image Set (' + str(img_set_size) + ' images)'
    ws['A2'] = 'Image Dir and File Name'
    ws['B2'] = 'Segmentation Dir and File Name'
    ws['C2'] = 'IoU Score'
    ws['D2'] = 'Pixel Accuracy'
    ws['E2'] = 'Is Day/Evening/Night Image'
    ws['F2'] = 'Seeds Location'
    ws['G2'] = 'Sky Pixel Accuracy'
    
    for col in ws.iter_cols(min_row=2, max_col=9, max_row=img_set_size+2):
        for cell in col:
            cell.alignment = pyxl.styles.Alignment(horizontal='center')
    
    img = cv2.imread(r"" + imgfile,0)
    img_color = cv2.imread(r"" + imgfile,1)
    isDay, seedList = getSeed(img_color)
    [nrow,ncol] = img.shape
    
    seedprint = ''
    seedprint += "Seed 1, (X : "+ str(seedList[0].row) + ", Y :" + str(seedList[0].col) +")"
    seedprint += "/Seed 2, (X : "+ str(seedList[1].row) + ", Y :" + str(seedList[1].col) +")"
    
    if not isDay:
        img = img * 2
    
    img = cv2.GaussianBlur(img,(5,5),3)
    
    regionMap = regionGrowing(img,seedList,2,'8_px')
    
    #Remove gap of ground objects and some sky cloud edges
    regionMap = cv2.dilate(regionMap,getConnect("8_np"),iterations=2)
    regionMap = cv2.erode(regionMap,getConnect("8_np"),iterations=2)
    
    #IoU calculation
    mask = cv2.imread(mask_dir,0)
    intersect = np.logical_and(regionMap, mask)
    union = np.logical_or(regionMap, mask)
    iou_score = np.sum(intersect) / np.sum(union)
    
    correct_sky_px = 0
    correct_ground_px = 0
    wrong_sky_px = 0
    wrong_ground_px = 0
    #Precision
    for row in range(nrow):
        for col in range(ncol):
            if mask[row,col] == 255 and regionMap[row,col] == 255:
                correct_sky_px += 1
            elif mask[row,col] == 0 and regionMap[row,col] == 0:
                correct_ground_px += 1
            elif mask[row,col] == 255 and regionMap[row,col] == 0:
                wrong_sky_px += 1
            elif mask[row,col] == 0 and regionMap[row,col] == 255:
                wrong_ground_px += 1
        
    pixel_accuracy = (correct_sky_px + correct_ground_px) / (correct_sky_px + correct_ground_px + wrong_sky_px + wrong_ground_px)
    sky_accuracy = correct_sky_px / (correct_sky_px + wrong_sky_px)
    #Final save processed image into segmentation result folder
    temp_img_name = imgfile
    segmented_img_name = temp_img_name.split("\\")
    segmented_img_name = segmented_img_name[1].split(".")
    segmented_img_name = segmented_img_name[0] + "_seg.jpg"
    segmented_folder_name = temp_img_name.split("\\")
    segmented_folder_name = segmented_folder_name[0]
    segmented_result_path = os.path.join("segmentation_result",segmented_folder_name)
    segmented_result_path = os.path.join(segmented_result_path, segmented_img_name)
    cv2.imwrite(segmented_result_path,regionMap)
    
    #Final save IoU score and other description for each segmentation
    cell_row += 1
    ws.cell(row=cell_row, column=1).value = imgfile
    ws.cell(row=cell_row, column=2).value = segmented_result_path
    ws.cell(row=cell_row, column=3).value = iou_score
    ws.cell(row=cell_row, column=4).value = pixel_accuracy
    ws.cell(row=cell_row, column=5).value = "Day/Evening" if isDay else "Night"
    ws.cell(row=cell_row, column=6).value = seedprint
    ws.cell(row=cell_row, column=7).value = sky_accuracy
    wb.save('iou_score//' +iou_file_name + '_iou_score.xlsx')

    #Printing segmentation Description
    img_count += 1
    print(str(img_count) + " " + iou_file_name + " Image done. Dir: " + imgfile + ", Mask Dir: " + mask_dir)
    day_print = "Day/Evening" if isDay else "Night"
    print("Image time: " + day_print)
    if iou_score > 0.7:
        print("IoU score : \033[1;32m" + str(iou_score) + "\033[1;0m")
    else:
        print("IoU score : \033[1;31m" + str(iou_score) + "\033[1;0m")
    if pixel_accuracy > 0.7:
        print("Pixel Accuracy : \033[1;32m" + str(pixel_accuracy))
    else:
        print("Pixel Accuracy : \033[1;31m" + str(pixel_accuracy))
    print("\033[1;0m--------------")
    
    #Plot to compare final segmented and the grey image parsed for region growing.
    pt.figure()
    pt.imshow(img)
    pt.show()
    
    pt.figure()
    pt.imshow(regionMap)
    pt.show()

#-----------------------------------------End----------------------------------------------------


